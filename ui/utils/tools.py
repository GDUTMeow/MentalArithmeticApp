import string
import random
import openpyxl
from io import BytesIO
from .app import judge, calculate_result
from typing import List, Dict


def generate_salt(length: int = 16) -> str:
    """
    @brief 生成哈希盐值

    @param length 哈希盐值长度
    @return str 盐值
    """
    return "".join(random.choices(string.ascii_letters + string.digits, k=16))


def generate_correct_answer_list(question_list: list) -> list[float]:
    """
    @brief 生成正确答案列表。

    @param question_list 问题列表
    @return list[float] 正确答案列表
    """
    return [question["correct_answer"] for question in question_list]


def c_strlen(s):
    """
    模拟C语言的strlen函数，计算字符串的长度。
    中文字符（C语言UTF-8编码下占3个字节）计为3，其他字符计为1。

    参数:
        s (str): 输入的字符串。

    返回:
        int: 按C语言算法计算的字符串长度。
    """
    length = 0
    for char in s:
        code_point = ord(char)
        # 判断是否为中文字符
        if (
            0x4E00 <= code_point <= 0x9FFF  # CJK统一汉字
            or 0x3400 <= code_point <= 0x4DBF  # CJK统一汉字扩展A
            or 0x20000 <= code_point <= 0x2A6DF  # CJK统一汉字扩展B
            or 0x2A700 <= code_point <= 0x2B73F  # CJK统一汉字扩展C
            or 0x2B740 <= code_point <= 0x2B81F  # CJK统一汉字扩展D
            or 0x2B820 <= code_point <= 0x2CEAF  # CJK统一汉字扩展E
            or 0xF900 <= code_point <= 0xFAFF  # CJK兼容汉字
            or 0x2F800 <= code_point <= 0x2FA1F  # CJK兼容汉字补充
        ):
            length += 3
        else:
            length += 1
    return length


def calculate_score(question_list: list, user_answer_list: list) -> int:
    """
    @brief 计算用户得分。

    @param question_list 问题列表
    @param user_answer_list 用户答案列表
    @return int 得分
    """
    # 计算总分
    right_count = 0
    for question, answer in zip(question_list, user_answer_list):
        if judge(
            calculate_result(
                question.get("num1"), question.get("num2"), question.get("op")
            ),
            answer,
        ):
            right_count += 1
    score = right_count/len(question_list) * 100
    
    return int(score)  # 分数只能是整数，否则数据库存储方面存储后会出问题


def questions_xlsx_parse(raw_data: bytes) -> list:
    """
    @brief 解析传入的xlsx的数据

    @param data 传入的二进制xlsx文件数据
    @return list 解析后的数据列表
    """
    try:
        # 将传入的二进制数据流转为ByteIO对象
        data = BytesIO(raw_data)
        # 打开Excel工作簿
        wb = openpyxl.load_workbook(data, data_only=True)
    except Exception as e:
        print(f"无法打开文件 {data}: {e}")
        return

    ws = wb.active

    results = []

    # 遍历每一行（跳过表头）
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            num1, operator, num2 = row[:3]

            # 检查是否为整数
            if (
                not isinstance(num1, int)
                or not isinstance(operator, int)
                or not isinstance(num2, int)
            ):
                print(f"第{row_idx}行：数据类型错误，跳过计算。")
                continue

            if operator == 3:
                if num2 == 0:
                    print(f"第{row_idx}行：除数不能为0，跳过计算。")
                    continue
            elif operator not in [0, 1, 2, 3]:
                print(f"第{row_idx}行：无效的操作符 {operator}，跳过计算。")
                continue
            operation = operator
            results.append((num1, operation, num2))

        except Exception as e:
            print(f"第{row_idx}行：发生错误: {e}，跳过计算。")
            continue

    return results


def students_xlsx_parser(raw_data: bytes) -> list:
    """
    @brief 解析传入的xlsx的数据

    @param data 传入的二进制xlsx文件数据
    @return list 解析后的数据列表
    """
    try:
        # 将传入的二进制数据流转为ByteIO对象
        data = BytesIO(raw_data)
        # 打开Excel工作簿
        wb = openpyxl.load_workbook(data, data_only=True)
    except Exception as e:
        print(f"无法打开文件 {data}: {e}")
        return

    ws = wb.active

    results = []

    # 遍历每一行（跳过表头）
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            number, name, student_class, password = row[:4]
            name = str(name)
            student_class = str(student_class)
            if not all([number, name, student_class, password]):
                break
            results.append((number, name, student_class, password))
        except Exception as e:
            print(f"第{row_idx}行：发生错误: {e}")
            continue
    return results

def generate_score_report(student_score_list: List[Dict]) -> BytesIO:
    """
    生成学生成绩报告的Excel二进制数据流。

    此函数接收一个包含学生成绩信息的列表，并将其导出为Excel格式的二进制数据流。
    Excel文件包含四列：学号、姓名、分数、逾期作答。

    参数:
        student_score_list (List[Dict]): 
            每个字典包含以下键：
                - "id" (str): 学生的学号
                - "name" (str): 学生的姓名
                - "score" (int): 学生的分数，未作答时为-1
                - "expired" (int): 逾期作答标记，1表示逾期，0表示未逾期，-1表示未作答

    返回:
        BytesIO: 包含生成的Excel文件数据的二进制流
    """
    
    # 创建一个新的Excel工作簿
    wb = openpyxl.Workbook()
    
    # 获取活动的工作表
    ws = wb.active
    ws.title = "学生成绩"  # 设置工作表名称为“学生成绩”

    # 定义表头并添加到工作表的第一行
    headers = ["学号", "姓名", "分数", "逾期作答"]
    ws.append(headers)

    # 遍历每个学生的成绩信息，并根据逻辑处理后添加到工作表
    for student in student_score_list:
        number = student["id"]         # 学号
        name = student["name"]         # 姓名
        score = student["score"]       # 分数
        expired = student["expired"]   # 逾期作答标记

        # 根据分数和逾期标记处理显示内容
        if score == -1:
            score_display = "未作答"          # 分数为-1时显示“未作答”
            expired_display = "未作答"        # 同时逾期标记也显示“未作答”
        else:
            score_display = score            # 存在分数时显示实际分数
            if expired == 1:
                expired_display = "是"        # 逾期作答标记为1时显示“是”
            elif expired == 0:
                expired_display = "否"        # 逾期作答标记为0时显示“否”
            else:
                expired_display = "未知"      # 处理其他可能的情况

        # 创建当前行的数据列表
        row = [number, name, score_display, expired_display]
        
        # 将当前行的数据添加到工作表中
        ws.append(row)

    # 将工作簿保存到二进制流中
    stream = BytesIO()
    wb.save(stream)          # 将Excel文件写入内存中的BytesIO对象
    stream.seek(0)           # 将指针移动到流的开头，以便后续读取

    return stream            # 返回包含Excel数据的二进制流

def is_chinese(char: str) -> bool:
    code_point = ord(char)
    if (
            0x4E00 <= code_point <= 0x9FFF  # CJK统一汉字
            or 0x3400 <= code_point <= 0x4DBF  # CJK统一汉字扩展A
            or 0x20000 <= code_point <= 0x2A6DF  # CJK统一汉字扩展B
            or 0x2A700 <= code_point <= 0x2B73F  # CJK统一汉字扩展C
            or 0x2B740 <= code_point <= 0x2B81F  # CJK统一汉字扩展D
            or 0x2B820 <= code_point <= 0x2CEAF  # CJK统一汉字扩展E
            or 0xF900 <= code_point <= 0xFAFF  # CJK兼容汉字
            or 0x2F800 <= code_point <= 0x2FA1F  # CJK兼容汉字补充
        ):
        return True
    else:
        return False